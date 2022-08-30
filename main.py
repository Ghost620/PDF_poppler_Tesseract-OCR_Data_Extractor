from selenium import webdriver
import undetected_chromedriver as uc
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from openpyxl import Workbook
import os, pytesseract, re
from pdf2image import convert_from_path
from openpyxl import Workbook
from multiprocessing.dummy import Pool as ThreadPool
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
from pathlib import Path
print(Path('C:', '/', 'Users'))
import sys
import shutil
import PySimpleGUI as sg
from datetime import date, time
import random,time
from datetime import datetime

dir_path = str(os.getcwd())
newpath = dir_path + "\PDF"
if not os.path.exists(newpath):
    os.makedirs(newpath)
else:
    shutil.rmtree(newpath)
    os.makedirs(newpath)

print(newpath)
options = webdriver.ChromeOptions()
options.add_experimental_option('prefs', {
"download.default_directory": newpath,
"download.prompt_for_download": False,
"download.directory_upgrade": True,
"plugins.always_open_pdf_externally": True
})


pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
driver = webdriver.Chrome(ChromeDriverManager().install(),options=options)
driver.get('https://search.jeffersondeeds.com/insttype.php')
driver.implicitly_wait(30)

element = driver.find_elements_by_xpath('//select[@name="itype1"]/option')
options_list = []

for i in element:
    options_list.append(i.text)
options_list.pop(0)

today = str(date.today()).split('-')
con = False
while True:
    theme_name_list = sg.theme_list()
    sg.theme(theme_name_list[random.randint(0, len(theme_name_list))])
    #define layout
    layout=[[sg.Text('Choose Instrument Type',size=(20, 1), font='Ubuntu',justification='left')],
            [sg.Combo(options_list,default_value=options_list[0],key='board')],
            [sg.Input(key='-IN4-', size=(20,1)), sg.CalendarButton('Calendar',font="Ubuntu",  target='-IN4-', default_date_m_d_y=(int(today[1]),int(today[2]),int(today[0])), )],
            [sg.Button('OK', font=('Ubuntu',12)),sg.Button('CANCEL', font=('Ubuntu',12))]]
    #Define Window
    win =sg.Window('Sir Ravi Application',layout)
    #Read  values entered by user
    e,v=win.read()
    print(e,v)
    if e == None or e == "CANCEL":
        print("exit")
        win.close()
        con = True
        driver.close()
        break
    else:
        if v['board'] == None or v['-IN4-'] == None or v['-IN4-'] == '' or v['board']=='':
            win.close()

            continue
        else:
            dates = v['-IN4-'].split(" ")[0].split('-')
            dates = dates[1]+'/'+dates[2]+'/'+dates[0]
            print(dates)
            instrument_type = v['board']
            print(instrument_type)
            win.close()
            con = False
            break

if con:
    pass
else:
    for i in driver.find_elements_by_tag_name('select'):
        Select(i).select_by_visible_text(instrument_type)
    try:
        WebDriverWait(driver,30).until(ec.element_to_be_clickable((By.ID,"datepickerbdate")))
    except:
        pass
    driver.find_element_by_id('datepickerbdate').clear()
    driver.find_element_by_id('datepickerbdate').send_keys(dates)

    driver.find_elements_by_tag_name('input')[-2].click()

    pdfs = [i.get_attribute('href') for i in driver.find_elements_by_xpath("//*[@id='selectOption']/span/a")]

    ##################################################################################################################
    for j in pdfs: 
        try:
            driver.get(j)
            
        except:
            driver.get(j)
            
            
    ##################################################################################################################
    # time.sleep(10)
    tries = 1
    while True:
        files = [f for f in os.listdir("./PDF")]
        files = list(filter(lambda f: f.endswith(('.pdf','.PDF')), files))

        if (len(files)==len(pdfs)):
            break
        else:
            time.sleep(3)
            tries += 1
            if (tries == 5):
                break

        

    # ####################################################################################################################

    search_list, zipcodes = [], []
    def threading_function(PDF_file):
        global search_list
       
        pages = convert_from_path(PDF_file, 500, poppler_path=r'C:\Program Files\poppler-0.68.0\bin')

        cond = False
        
        for page in pages:
        
            try:
                if cond == True:
                    break

                text = str(((pytesseract.image_to_string(page))))
                for index, i in enumerate(text.split(',')):
                    ZipCode = '-'
                    if 'property address: ' in i.lower() or 'property address; ' in i.lower() or 'address: ' in i.lower():
                        ind = re.search(r'address', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break

                    elif 'unknown occupants of' in  i.lower():
                        ind = re.search(r'occupants of ', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break

                    elif 'involved or affected is of' in  i.lower():
                        ind = re.search(r'affected is of ', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break
                    elif 'commonly known as' in  i.lower():
                        ind = re.search(r'commonly known as ', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break
                    elif 'property known as' in  i.lower():
                        ind = re.search(r'known as ', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break
                    elif 'referred to as' in  i.lower():
                        ind = re.search(r'referred to as ', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break
                    elif 'located at' in  i.lower():
                        ind = re.search(r'located at ', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break
                    elif 'address of' in  i.lower():
                        ind = re.search(r'address of ', i.lower())
                        cond = True
                        if ind:
                            ind = ind.span()[1]
                            search = i[ind:]
                            print(search)
                            search_list.append(search)
                            for j in range(1, 4):                    
                                zip_code = re.search('(\d{5})([- ])?(\d{4})?', text.split(',')[index+j])
                                if zip_code != None:
                                    ZipCode = zip_code.groups()[0]
                            zipcodes.append(ZipCode)
                            print(ZipCode)
                            break
                
            except Exception as e:
                print(e)
                print(ZipCode)
                zipcodes.append(ZipCode)

    pool = ThreadPool(4)

    for k in files:
        PDF_file = k
        print(f'{newpath}\{PDF_file}')
        pool.apply_async(threading_function, (f'{newpath}\{PDF_file}',))

    pool.close()
    pool.join()
    driver.close()
                    
    ####################################################################################################################

    for i in range(len(search_list)):
        if ': ' in search_list[i]:
            search_list[i] = search_list[i].replace(': ', '')
        if '; ' in search_list[i]:
            search_list[i] = search_list[i].replace('; ', '')
        if '\n' in search_list[i]:
            search_list[i] = search_list[i].replace('\n', ' ')
        if '"' in search_list[i]:
            search_list[i] = search_list[i].replace('"', '')
        if '™' in search_list[i]:
            search_list[i] = search_list[i].replace('™', '')
        if ' a/k/a ' in search_list[i]:
            search_list[i] = search_list[i].split('a/k/a')[1]
        if 'South' in search_list[i]:
            search_list[i] = search_list[i].replace('South', 'S')
        if 'North' in search_list[i]:
            search_list[i] = search_list[i].replace('North', 'N')
        if 'Street' in search_list[i]:
            search_list[i] = search_list[i].replace('Street', 'ST')
        if 'Avenue' in search_list[i]:
            search_list[i] = search_list[i].replace('avenue', 'AVE')
        if 'Avenuc' in search_list[i]:
            search_list[i] = search_list[i].replace('avenuc', 'AVE')
        if '| ' in search_list[i]:
            search_list[i] = search_list[i].replace('| ', '')
        if '!' in search_list[i]:
            search_list[i] = search_list[i].replace('!', '')
        if 'or near ' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace('or near ', '')
        if 'and known as ' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace('and known as ', '')
        if ' §' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace(' §', '')
        if 'both ' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace('both ', '')
        if '..' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace('..', '')
        if '.' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace('.', '')
        if ' louisville' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace(' louisville', '')
        if '  ' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace('  ', ' ')
        if '(' in search_list[i].lower():
            search_list[i] = search_list[i].split('(')[0]
        if '&' in search_list[i].lower():
            search_list[i] = search_list[i].split(' & ')[-1]
        if '#' in search_list[i].lower():
            search_list[i] = search_list[i].lower().replace('#', 'APT ')
        if ' — ' in search_list[i].lower():
            search_list[i] = search_list[i].split(' — ')[-1]
        
    # ####################################################################################################################        
    owners, values, addresses, ZIP = [], [], [], []


    def thread_function(ind,j):
        driver = webdriver.Chrome(ChromeDriverManager().install())        
        driver.get('https://jeffersonpva.ky.gov/property-search/')
        driver.implicitly_wait(1)
        try:
            driver.find_element_by_id('psfldAddress').clear()
            driver.find_element_by_id('psfldAddress').send_keys(j)
            driver.find_element_by_id('psfldAddress').send_keys(Keys.ENTER)

            try:
                dt = WebDriverWait(driver, 2).until(ec.visibility_of_element_located((By.TAG_NAME, "dt")))
            except:
                print(j)
                try:
                    WebDriverWait(driver, 2).until(ec.visibility_of_element_located((By.CLASS_NAME, "suggestion"))).click()
                except:
                    pass
                if driver.find_elements_by_tag_name('h3')[3].text == '0 records found':
                    driver.back()
                    halfj = j.split(' ')[0] + ' ' + j.split(' ')[1] + ' ' + j.split(' ')[2]
                    driver.find_element_by_id('psfldAddress').clear()
                    driver.find_element_by_id('psfldAddress').send_keys(halfj)
                    driver.find_element_by_id('psfldAddress').send_keys(Keys.ENTER)

                try:
                    for i in driver.find_elements_by_xpath("//*[@id='content']/table/tbody/tr/td[2]/a"):
                        if j.split(' ')[-1] in i.text.lower():
                            driver.get(i.get_attribute('href'))
                            break
                    else:
                        driver.get(driver.find_element_by_xpath("//*[@id='content']/table/tbody/tr/td[2]/a").get_attribute('href'))
                except:
                    pass

            for i in range(len(driver.find_elements_by_tag_name('dt'))):
                if driver.find_elements_by_tag_name('dt')[i].text == 'Owner':
                    owners.append(driver.find_elements_by_tag_name('dd')[i].text)

                    for k in driver.find_elements_by_tag_name('h1'):
                        if k.text == '':
                            continue
                        property_address = k.text
                        print(property_address)
                    ZIP.append(zipcodes[ind])

                    addresses.append(property_address)
                    print(f"OWNER: {driver.find_elements_by_tag_name('dd')[i].text}")
                elif driver.find_elements_by_tag_name('dt')[i].text == 'Assessed Value':
                    values.append(driver.find_elements_by_tag_name('dd')[i].text)
                    print(f"VALUE: {driver.find_elements_by_tag_name('dd')[i].text}")
        except Exception as e:
            print(e)
            print(j) 
        driver.close()
    pool = ThreadPool(4)
    for ind, j in enumerate(search_list):
        pool.apply_async(thread_function, (ind,j,))
    pool.close()
    pool.join()


    values = [int(i.replace(',', '')) for i in values]

    workbook = Workbook()
    sheet = workbook.active
    bold_font = Font(bold=True,size = "15")

    center_aligned_text = Alignment(horizontal="center")
    double_border_side = Side(border_style="double")
    square_border = Border(top=double_border_side,
                    right=double_border_side,
                    bottom=double_border_side,
                    left=double_border_side)
    sheet["A1"] = "First Name"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Last Name"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Property Address"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Zip Code"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Assessed Value"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border

    for i in range(len(owners)):
        sheet[f"A{i+2}"] = ' '.join(map(str, owners[i].split(' ')[1:]))
        sheet[f"B{i+2}"] = owners[i].split(' ')[0]
        sheet[f"C{i+2}"] = addresses[i]
        sheet[f"D{i+2}"] = ZIP[i]
        sheet[f"E{i+2}"] = values[i]

    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder
    sheet.freeze_panes = "B1"
    workbook.save(filename=f"{instrument_type}[{dates.replace('/','-')} - {datetime.today().strftime('%d-%m-%Y')}].xlsx")
        
    # ####################################################################################################################
    def captcha_solver(driver1):
        try:
            while True:
                element = driver1.find_element_by_css_selector('#px-captcha')
                action = ActionChains(driver1)
                click = ActionChains(driver1)
                action.click_and_hold(element)
                action.perform()
                time.sleep(10)
                action.release(element)
                action.perform()
                time.sleep(0.2)
                action.release(element)
        except Exception as e:
            print(e)
            return None
    driver1 = uc.Chrome()
    driver1.get('https://www.zillow.com/homes/_rb/')
    driver1.implicitly_wait(1)
    captcha_solver(driver1)


    lst = []

    for a in addresses:
        bed, bath, sqft, status, zes, date, event, price = '-', '-', '-', '-', '-', '-', '-', '-'
        try:
            inp = driver1.find_element_by_tag_name('input')
            inp.send_keys(Keys.CONTROL, 'A')
            inp.send_keys(Keys.DELETE)
            inp.send_keys(a)
            inp.send_keys(Keys.ENTER)
            captcha_solver(driver1)

            try:
                details = WebDriverWait(driver1, 2).until(ec.visibility_of_element_located((By.CLASS_NAME, "summary-container")))
            except Exception as e:
                print(e)
                details = WebDriverWait(driver1, 2).until(ec.visibility_of_element_located((By.CLASS_NAME, "ds-home-details-chip")))


            for i in details.text.split('\n'):
                if 'bd' in i:
                    bed = i.split('bd')[0][-2]
                    bath = i.split('bd')[-1].split(' ')[0]
                    sqft = i.split('ba')[-1].split(' ')[0]
                if 'Zestimate®:' in i:
                    status, zes = i.split('Zestimate®: ')[0], i.split('Zestimate®: ')[1]
                    if 'None' in zes:
                        zes = 'None'
                    if 'Rent' in zes:
                        zes = zes.replace('Rent', '')
                    if 'assessed' in zes:
                        zes = zes.split(' ')[0]
                    if 'Sold' in status:
                        status = 'Sold'


            try:    
                try:
                    driver1.find_element_by_link_text('Neighborhood details').click()
                except Exception as e:
                    print(e)
                    driver1.find_element_by_link_text('Price and tax history').click()
                price_history = driver1.find_elements_by_tag_name('tr')[1].text.split(' ')
                date, event, price = price_history[0], price_history[1], price_history[2]
            except Exception as e:
                print(e)
                pass


            driver1.back()
            print(f"""ADDRESS: {a}
            Bed: {bed}
            Bath: {bath}
            Area: {sqft}
            Status: {status}
            Zestimate: {zes}
            Date: {date}
            Event: {event}
            Price: {price}
            """)
        except Exception as e:
            print(e)
            try:
                WebDriverWait(driver1, 2).until(ec.visibility_of_element_located(driver1.find_element_by_css_selector('#px-captcha')))
                captcha_solver(driver1)
            except Exception as e:
                print(e)
                driver1.get('https://www.zillow.com/homes/_rb/')
                captcha_solver(driver1)


        lst.append({'address': a,
        'bed': bed,
        'bath': bath,
        'area': sqft,
        'status': status,
        'zestimate': zes,
        'date': date,
        'event': event,
        'price': price})    


    workbook = Workbook()
    sheet = workbook.active

    sheet["A1"] = "Address"
    sheet["A1"].font = bold_font
    sheet["A1"].alignment = center_aligned_text
    sheet["A1"].border = square_border
    sheet["B1"] = "Bed"
    sheet["B1"].font = bold_font
    sheet["B1"].alignment = center_aligned_text
    sheet["B1"].border = square_border
    sheet["C1"] = "Bath"
    sheet["C1"].font = bold_font
    sheet["C1"].alignment = center_aligned_text
    sheet["C1"].border = square_border
    sheet["D1"] = "Area"
    sheet["D1"].font = bold_font
    sheet["D1"].alignment = center_aligned_text
    sheet["D1"].border = square_border
    sheet["E1"] = "Status"
    sheet["E1"].font = bold_font
    sheet["E1"].alignment = center_aligned_text
    sheet["E1"].border = square_border
    sheet["F1"] = "Zestimate"
    sheet["F1"].font = bold_font
    sheet["F1"].alignment = center_aligned_text
    sheet["F1"].border = square_border
    sheet["G1"] = "Date"
    sheet["G1"].font = bold_font
    sheet["G1"].alignment = center_aligned_text
    sheet["G1"].border = square_border
    sheet["H1"] = "Event"
    sheet["H1"].font = bold_font
    sheet["H1"].alignment = center_aligned_text
    sheet["H1"].border = square_border
    sheet["I1"] = "Price"
    sheet["I1"].font = bold_font
    sheet["I1"].alignment = center_aligned_text
    sheet["I1"].border = square_border

    for i in range(len(lst)):
        sheet[f"A{i+2}"] = lst[i]['address']
        sheet[f"B{i+2}"] = lst[i]['bed']
        sheet[f"C{i+2}"] = lst[i]['bath']
        sheet[f"D{i+2}"] = lst[i]['area'] + ' sqft'
        sheet[f"E{i+2}"] = lst[i]['status']
        sheet[f"F{i+2}"] = lst[i]['zestimate']
        sheet[f"G{i+2}"] = lst[i]['date']
        sheet[f"H{i+2}"] = lst[i]['event']
        sheet[f"I{i+2}"] = lst[i]['price']
    dim_holder = DimensionHolder(worksheet=sheet)

    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)

    sheet.column_dimensions = dim_holder

    workbook.save(filename=f"Zillow[{dates.replace('/','-')} - {datetime.today().strftime('%d-%m-%Y')}].xlsx")
